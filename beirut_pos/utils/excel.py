# beirut_pos/services/excel.py
"""Lightweight XLSX writer used for exported reports.

The previous implementation built the XLSX archive manually and locked sheets
with a password. That approach produced files that many Excel viewers (notably
on Windows) flagged as corrupted or required a password to open. The helper now
uses :mod:`openpyxl` to emit standard, editable workbooks without any
protection so reports open cleanly everywhere.
"""

from __future__ import annotations

from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from xml.sax.saxutils import escape
import re

# XML 1.0 forbids control chars except TAB(0x09), LF(0x0A), CR(0x0D)
_XML_CTRL_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def _sanitize_xml_text(value) -> str:
    """Remove illegal XML control chars and escape for XML."""
    s = "" if value is None else str(value)
    s = _XML_CTRL_RE.sub("", s)  # strip disallowed control characters
    return escape(s).strip()


def _column_letter(index: int) -> str:
    """Convert 0-based column index to Excel letter (0->A, 25->Z, 26->AA)."""
    # openpyxl uses 1-based indexing; reuse helper for any legacy calls.
    return get_column_letter(index + 1)


def _sanitize_sheet_name(name: str) -> str:
    """Clean sheet name to be Excel-compatible."""
    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if ch in invalid else ch for ch in name.strip())
    cleaned = cleaned or "Report"
    return cleaned[:31]


def write_protected_workbook(
    path: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[str]],
    *,
    title: str = "Report",
    password: str | None = None,
) -> None:
    """Create an editable XLSX file with the given data.

    Args:
        path: Output file path (should end with .xlsx)
        headers: Column headers
        rows: Data rows (list of lists)
        title: Report title (used as sheet name)
        password: Deprecated. Kept for backward compatibility; no protection is
            applied.
    """

    sheet_name = _sanitize_sheet_name(title)

    # Sanitize all cell values to prevent XML/Excel issues
    header_values = [_sanitize_xml_text(h) for h in headers]
    table_rows = [[_sanitize_xml_text(cell) for cell in row] for row in rows]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers and rows
    ws.append(header_values)
    for row in table_rows:
        ws.append(row)

    # Auto-fit column widths based on longest cell (rough approximation)
    for col_idx, header in enumerate(header_values, start=1):
        values = [header] + [row[col_idx - 1] if len(row) >= col_idx else "" for row in table_rows]
        max_len = max(len(str(value)) for value in values) if values else 0
        if max_len:
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    try:
        wb.save(path)
    except PermissionError:
        raise Exception("الملف مفتوح في برنامج آخر. أغلقه وحاول مرة أخرى.")
    except Exception as e:
        raise Exception(f"خطأ في إنشاء ملف Excel: {str(e)}")
