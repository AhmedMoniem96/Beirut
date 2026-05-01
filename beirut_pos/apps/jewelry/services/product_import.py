from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

from .barcode_printer import _SUPPORTED_BARCODE_TYPES
from .db import upsert_product_by_sku
from beirut_pos.utils.excel import write_protected_workbook


@dataclass
class ImportRowResult:
    row_number: int
    sku: str
    status: str
    message: str = ""


@dataclass
class ProductImportReport:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0
    rows: list[ImportRowResult] = field(default_factory=list)


REQUIRED_HEADERS = [
    "name_ar",
    "name_en",
    "sku",
    "barcode",
    "barcode_type",
    "price",
    "qty_on_hand",
]

OPTIONAL_HEADERS = ["min_qty", "category", "handmade_flag", "stone_type", "color"]


class ProductImportError(ValueError):
    pass


def _as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _as_float(value: Any, field_name: str) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ProductImportError(f"{field_name} must be numeric") from exc


def _as_bool(value: Any) -> bool:
    txt = _as_text(value).lower()
    return txt in {"1", "true", "yes", "y", "on"}


def _validate_barcode_type(raw: str) -> str:
    normalized = raw.strip().lower().replace(" ", "").replace("-", "")
    if normalized == "qrcode":
        normalized = "qr"
    if not normalized:
        return ""
    if normalized not in _SUPPORTED_BARCODE_TYPES:
        supported = ", ".join(_SUPPORTED_BARCODE_TYPES.values())
        raise ProductImportError(f"Invalid barcode_type. Supported: {supported}")
    return _SUPPORTED_BARCODE_TYPES[normalized]


def _parse_headers(sheet) -> dict[str, int]:
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = {_as_text(value).lower(): idx for idx, value in enumerate(header_cells)}
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ProductImportError(f"Missing required headers: {', '.join(missing)}")
    return headers


def import_products_from_excel(path: str) -> ProductImportReport:
    workbook = load_workbook(filename=path, data_only=True)
    sheet = workbook.active
    headers = _parse_headers(sheet)
    report = ProductImportReport()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        data = {key: row[index] if index < len(row) else None for key, index in headers.items()}
        sku = _as_text(data.get("sku"))
        try:
            name_ar = _as_text(data.get("name_ar"))
            name_en = _as_text(data.get("name_en"))
            if not name_ar or not name_en or not sku:
                raise ProductImportError("name_ar, name_en, and sku are required")
            payload = {
                "name_ar": name_ar,
                "name_en": name_en,
                "sku": sku,
                "barcode": _as_text(data.get("barcode")),
                "barcode_type": _validate_barcode_type(_as_text(data.get("barcode_type"))),
                "price": _as_float(data.get("price"), "price"),
                "qty_on_hand": _as_float(data.get("qty_on_hand"), "qty_on_hand"),
                "min_qty": _as_float(data.get("min_qty"), "min_qty"),
                "category": _as_text(data.get("category")),
                "handmade_flag": _as_bool(data.get("handmade_flag")),
                "stone_type": _as_text(data.get("stone_type")),
                "color": _as_text(data.get("color")),
            }
            outcome = upsert_product_by_sku(payload)
            if outcome == "created":
                report.created += 1
            elif outcome == "updated":
                report.updated += 1
            else:
                report.skipped += 1
            report.rows.append(ImportRowResult(row_number=row_number, sku=sku, status=outcome))
        except ProductImportError as exc:
            report.errors += 1
            report.rows.append(ImportRowResult(row_number=row_number, sku=sku, status="error", message=str(exc)))

    return report


def generate_import_template(path: str) -> None:
    headers = REQUIRED_HEADERS + OPTIONAL_HEADERS
    write_protected_workbook(path, headers, [], title="Product Import Template")
